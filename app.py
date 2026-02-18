import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook

st.set_page_config(page_title="Daily Sales Dashboard", layout="wide")

st.title("📊 Daily Sales Dashboard")

# ---------------------------
# Upload Excel File
# ---------------------------
uploaded_file = st.file_uploader("Upload Monthly Excel File", type=["xlsx"])

if uploaded_file:

    wb = load_workbook(uploaded_file, data_only=True)
    sheets = wb.sheetnames

    selected_sheet = st.selectbox("Select Date Sheet", sheets)

    ws = wb[selected_sheet]

    # Extract CONSOLIDATE DATA section
    data = ws.iter_rows(min_row=7, max_row=14, min_col=27, max_col=35, values_only=True)

    df = pd.DataFrame(data)

    # Assign correct headers
    df.columns = [
        "SHIFT",
        "QTY",
        "SALE_AMOUNT",
        "CASH",
        "PAYTM",
        "ATM",
        "CREDIT_SALE",
        "TOTAL_COLLECTION",
        "DIFF"
    ]

    # Remove header row (SHIFT, QTY...)
    df = df.iloc[1:]

    # Clean numeric columns
    numeric_cols = df.columns[1:]

    for col in numeric_cols:
        df[col] = (
            df[col]
            .astype(str)
            .str.replace(",", "", regex=False)
            .str.strip()
        )
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Split shift rows and total row
    shift_df = df.iloc[:3]
    total_row = df.iloc[3]

    # ---------------------------
    # TOTAL METRICS
    # ---------------------------
    total_qty = total_row["QTY"]
    total_atm = total_row["ATM"]
    total_collection = total_row["TOTAL_COLLECTION"]
    total_diff = total_row["DIFF"]

    # ---------------------------
    # METRICS UI
    # ---------------------------
    st.subheader("🔢 Overall Totals")

    col1, col2, col3, col4 = st.columns(4)

    col1.metric("Total Quantity", f"{total_qty:,.2f}")
    col2.metric("Total ATM", f"{total_atm:,.2f}")
    col3.metric("Total Collection", f"{total_collection:,.2f}")
    col4.metric("Difference", f"{total_diff:,.2f}")

    st.divider()

    # ---------------------------
    # SHIFT TABLE
    # ---------------------------
    st.subheader("📋 Shift Wise Data")

    st.dataframe(
        shift_df.style.format({
            "QTY": "{:,.2f}",
            "SALE_AMOUNT": "{:,.2f}",
            "CASH": "{:,.2f}",
            "PAYTM": "{:,.2f}",
            "ATM": "{:,.2f}",
            "CREDIT_SALE": "{:,.2f}",
            "TOTAL_COLLECTION": "{:,.2f}",
            "DIFF": "{:,.2f}",
        }),
        use_container_width=True
    )

    # ---------------------------
    # CHART 1 - Sale Amount by Shift
    # ---------------------------
    st.subheader("📊 Sale Amount by Shift")

    fig1 = px.bar(
        shift_df,
        x="SHIFT",
        y="SALE_AMOUNT",
        text="SALE_AMOUNT",
        color="SHIFT"
    )

    fig1.update_layout(height=400)
    st.plotly_chart(fig1, use_container_width=True)

    # ---------------------------
    # CHART 2 - Payment Distribution
    # ---------------------------
    st.subheader("💳 Payment Mode Distribution")

    payment_df = shift_df[["SHIFT", "CASH", "PAYTM", "CREDIT_SALE"]]

    payment_melted = payment_df.melt(
        id_vars="SHIFT",
        var_name="MODE",
        value_name="AMOUNT"
    )

    fig2 = px.bar(
        payment_melted,
        x="SHIFT",
        y="AMOUNT",
        color="MODE",
        barmode="group"
    )

    fig2.update_layout(height=400)
    st.plotly_chart(fig2, use_container_width=True)

else:
    st.info("Please upload your monthly Excel file to begin.")
